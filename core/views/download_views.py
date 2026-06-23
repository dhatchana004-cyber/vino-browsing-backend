"""
Data Download API views — Excel and CSV exports using openpyxl.
"""
import io
from datetime import date
from decimal import Decimal
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ..models import ServiceEntry, Customer, Expense
from ..permissions import IsOwner


def style_worksheet(ws, headers):
    """Apply consistent professional styling to worksheet."""
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='0F1729', end_color='0F1729', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Set column widths
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = 18

    ws.auto_filter.ref = ws.dimensions


def add_entry_rows(ws, entries, start_row=2):
    """Add entry data rows with alternating colors."""
    alt_fill = PatternFill(start_color='F0F4FF', end_color='F0F4FF', fill_type='solid')

    for idx, entry in enumerate(entries):
        row = start_row + idx
        row_data = [
            idx + 1,
            entry.date.strftime('%Y-%m-%d'),
            entry.customer_name or (entry.customer.name if entry.customer else ''),
            entry.phone,
            entry.service.name,
            entry.srn_number,
            float(entry.amount),
            float(entry.charge),
            float(entry.profit),
            entry.get_status_display(),
            entry.staff.get_full_name() or entry.staff.username,
            entry.remarks,
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if idx % 2 == 1:
                cell.fill = alt_fill

    return start_row + len(entries)


def add_summary_row(ws, row, entries, total_expenses=None):
    """Add totals summary row at the bottom."""
    totals = entries.aggregate(
        total_amount=Sum('amount'),
        total_charge=Sum('charge'),
        total_profit=Sum('profit'),
    )
    summary_font = Font(bold=True, size=11)
    summary_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
    expense_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
    profit_fill = PatternFill(start_color='D1ECF1', end_color='D1ECF1', fill_type='solid')

    ws.cell(row=row, column=5, value='TOTALS').font = summary_font
    for col, key in [(7, 'total_amount'), (8, 'total_charge'), (9, 'total_profit')]:
        cell = ws.cell(row=row, column=col, value=float(totals[key] or 0))
        cell.font = summary_font
        cell.fill = summary_fill

    if total_expenses is not None:
        row += 1
        ws.cell(row=row, column=5, value='TOTAL EXPENSES').font = summary_font
        cell = ws.cell(row=row, column=7, value=float(total_expenses))
        cell.font = summary_font
        cell.fill = expense_fill
        
        row += 1
        final_profit = float(totals['total_profit'] or 0) - float(total_expenses)
        ws.cell(row=row, column=5, value='FINAL PROFIT (Profit - Expenses)').font = summary_font
        cell = ws.cell(row=row, column=7, value=final_profit)
        cell.font = summary_font
        cell.fill = profit_fill


ENTRY_HEADERS = [
    '#', 'Date', 'Customer', 'Phone', 'Service', 'SRN',
    'Amount (₹)', 'Charge (₹)', 'Profit (₹)', 'Status', 'Staff', 'Remarks'
]


class DownloadAllView(APIView):
    """GET /api/download/all/ — All records Excel."""
    permission_classes = [IsOwner]

    def get(self, request):
        entries = ServiceEntry.objects.select_related('service', 'staff', 'customer').order_by('-date')
        total_exp = Expense.objects.aggregate(t=Sum('amount'))['t'] or Decimal('0')
        wb = Workbook()
        ws = wb.active
        ws.title = 'All Records'
        style_worksheet(ws, ENTRY_HEADERS)
        last_row = add_entry_rows(ws, entries)
        add_summary_row(ws, last_row, entries, total_exp)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=VINO_All_Records.xlsx'
        wb.save(response)
        return response


class DownloadDailyView(APIView):
    """GET /api/download/daily/?start_date=YYYY-MM-DD&end_date=YYYY-MM-DD"""
    permission_classes = [IsOwner]

    def get(self, request):
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        entries = ServiceEntry.objects.select_related('service', 'staff', 'customer')
        
        if start_date and end_date:
            entries = entries.filter(date__range=[start_date, end_date])
            expenses = Expense.objects.filter(date__range=[start_date, end_date])
        else:
            today = timezone.localdate()
            entries = entries.filter(date=today)
            expenses = Expense.objects.filter(date=today)

        entries = entries.order_by('-created_at')
        total_exp = expenses.aggregate(t=Sum('amount'))['t'] or Decimal('0')

        wb = Workbook()
        ws = wb.active
        ws.title = f"{start_date} to {end_date}" if start_date else "Daily"
        style_worksheet(ws, ENTRY_HEADERS)
        last_row = add_entry_rows(ws, entries)
        add_summary_row(ws, last_row, entries, total_exp)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'VINO_Records_{start_date}_to_{end_date}.xlsx' if start_date else 'VINO_Records_Daily.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response


class DownloadMonthlyView(APIView):
    """GET /api/download/monthly/?month=MM&year=YYYY"""
    permission_classes = [IsOwner]

    def get(self, request):
        now = timezone.localdate()
        month = int(request.query_params.get('month', now.month))
        year = int(request.query_params.get('year', now.year))

        entries = ServiceEntry.objects.filter(
            date__year=year, date__month=month
        ).select_related('service', 'staff', 'customer').order_by('-date')
        
        expenses = Expense.objects.filter(date__year=year, date__month=month)
        total_exp = expenses.aggregate(t=Sum('amount'))['t'] or Decimal('0')

        wb = Workbook()
        ws = wb.active
        ws.title = f'{year}-{month:02d}'
        style_worksheet(ws, ENTRY_HEADERS)
        last_row = add_entry_rows(ws, entries)
        add_summary_row(ws, last_row, entries, total_exp)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'VINO_Records_{year}-{month:02d}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response


class DownloadYearlyView(APIView):
    """GET /api/download/yearly/?year=YYYY"""
    permission_classes = [IsOwner]

    def get(self, request):
        year = int(request.query_params.get('year', timezone.localdate().year))

        entries = ServiceEntry.objects.filter(
            date__year=year
        ).select_related('service', 'staff', 'customer').order_by('-date')
        
        expenses = Expense.objects.filter(date__year=year)
        total_exp = expenses.aggregate(t=Sum('amount'))['t'] or Decimal('0')

        wb = Workbook()
        ws = wb.active
        ws.title = str(year)
        style_worksheet(ws, ENTRY_HEADERS)
        last_row = add_entry_rows(ws, entries)
        add_summary_row(ws, last_row, entries, total_exp)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'VINO_Records_{year}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response


class DownloadStaffView(APIView):
    """GET /api/download/staff/?staff_id=ID"""
    permission_classes = [IsOwner]

    def get(self, request):
        from ..models import Attendance
        staff_id = request.query_params.get('staff_id')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        entries = Attendance.objects.select_related('staff')
        if staff_id:
            entries = entries.filter(staff_id=staff_id)
        if start_date:
            entries = entries.filter(date__gte=start_date)
        if end_date:
            entries = entries.filter(date__lte=end_date)
            
        entries = entries.order_by('-date', 'staff__first_name', 'login_time')

        first_login_map = {}
        for entry in entries:
            key = (entry.date, entry.staff_id)
            if key not in first_login_map:
                first_login_map[key] = entry.id
        first_login_ids = set(first_login_map.values())

        wb = Workbook()
        ws = wb.active
        ws.title = 'Staff Attendance'
        headers = ['#', 'Date', 'Staff', 'Login Time', 'Logout Time', 'Working Hours']
        style_worksheet(ws, headers)

        alt_fill = PatternFill(start_color='F0F4FF', end_color='F0F4FF', fill_type='solid')
        for idx, entry in enumerate(entries):
            row = 2 + idx
            local_logout = timezone.localtime(entry.logout_time) if entry.logout_time else None
            logout = local_logout.strftime('%I:%M %p') if local_logout else '-'
            
            local_login = timezone.localtime(entry.login_time)
            is_first_login = entry.id in first_login_ids
            is_late = is_first_login and (local_login.hour > 10 or (local_login.hour == 10 and local_login.minute > 0))
            login_str = local_login.strftime('%I:%M %p') + (' (LATE)' if is_late else '')
            
            if entry.working_hours:
                total_seconds = int(entry.working_hours.total_seconds())
                hours, remainder = divmod(total_seconds, 3600)
                minutes, _ = divmod(remainder, 60)
                wh_str = f"{hours}h {minutes}m"
            else:
                wh_str = '-'

            row_data = [
                idx + 1,
                entry.date.strftime('%Y-%m-%d'),
                entry.staff.get_full_name() or entry.staff.username,
                login_str,
                logout,
                wh_str
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                if idx % 2 == 1:
                    cell.fill = alt_fill

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=VINO_Staff_Attendance.xlsx'
        wb.save(response)
        return response


class DownloadCustomersView(APIView):
    """GET /api/download/customers/ — CSV for Gmail import."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        import csv
        
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        customers = Customer.objects.all().order_by('name')
        
        if start_date:
            customers = customers.filter(created_at__date__gte=start_date)
        if end_date:
            customers = customers.filter(created_at__date__lte=end_date)
            
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=VINO_Customers.csv'

        writer = csv.writer(response)
        writer.writerow(['Name', 'Phone', 'Address', 'Notes'])
        for c in customers:
            writer.writerow([c.name, c.phone, c.address, c.notes])

        return response


class DownloadExpensesView(APIView):
    """GET /api/download/expenses/ — Excel export for expenses."""
    permission_classes = [IsOwner]

    def get(self, request):
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        expenses = Expense.objects.select_related('created_by').order_by('-date', '-created_at')
        if start_date and end_date:
            expenses = expenses.filter(date__range=[start_date, end_date])

        wb = Workbook()
        ws = wb.active
        ws.title = 'Expenses'
        headers = ['#', 'Date', 'Title', 'Amount (₹)', 'Staff', 'Added On']
        style_worksheet(ws, headers)

        alt_fill = PatternFill(start_color='F0F4FF', end_color='F0F4FF', fill_type='solid')
        for idx, exp in enumerate(expenses):
            row = 2 + idx
            row_data = [
                idx + 1,
                exp.date.strftime('%Y-%m-%d'),
                exp.title,
                float(exp.amount),
                exp.created_by.get_full_name() or exp.created_by.username if exp.created_by else '-',
                exp.created_at.strftime('%Y-%m-%d %H:%M')
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                if idx % 2 == 1:
                    cell.fill = alt_fill

        # Summary
        total = expenses.aggregate(t=Sum('amount'))['t'] or 0
        summary_font = Font(bold=True, size=11)
        summary_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
        last_row = 2 + len(expenses)
        ws.cell(row=last_row, column=3, value='TOTAL EXPENSES').font = summary_font
        cell = ws.cell(row=last_row, column=4, value=float(total))
        cell.font = summary_font
        cell.fill = summary_fill

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=VINO_Expenses.xlsx'
        wb.save(response)
        return response
